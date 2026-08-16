"""从 CSV 或 Excel 文件读取测试用例数据。"""

from __future__ import annotations

import csv
import os
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook


TRUE_VALUES = {"1", "true", "yes", "y", "是", "勾选"}
FALSE_VALUES = {"0", "false", "no", "n", "否", "不勾选"}
SUPPORTED_DATA_SOURCES = {"auto", "file", "code"}


def parse_bool(value: Any, default: bool = False) -> bool:
    """把 Excel/CSV 中的常见布尔写法转换为 Python bool。"""
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(value)
    text = str(value or "").strip().lower()
    if text in TRUE_VALUES:
        return True
    if text in FALSE_VALUES:
        return False
    return default


def _normalize_value(key: str, value: Any) -> Any:
    """统一外部文件单元格格式，同时保留布尔列语义。"""
    if key == "agreement" or key == "execute":
        return parse_bool(value, default=True)
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return str(value).strip()


def _is_blank_row(row: Iterable[Any]) -> bool:
    """跳过全空行，避免把 Excel 或 CSV 末尾空行读成用例。"""
    return not any(str(value or "").strip() for value in row)


def read_csv_cases(path: str | Path) -> list[dict[str, Any]]:
    """读取 CSV 文件，支持带 BOM 的 utf-8-sig 文件。"""
    csv_path = Path(path)
    if not csv_path.exists():
        raise FileNotFoundError(f"测试数据文件不存在: {csv_path}")

    with csv_path.open("r", encoding="utf-8-sig", newline="") as file:
        reader = csv.DictReader(file)
        headers = [header.strip() for header in (reader.fieldnames or [])]
        cases: list[dict[str, Any]] = []
        for row in reader:
            if _is_blank_row(row.values()):
                continue
            cases.append(
                {
                    header: _normalize_value(header, row.get(header))
                    for header in headers
                }
            )
    return cases


def read_excel_cases(path: str | Path) -> list[dict[str, Any]]:
    """读取 Excel 第一个工作表的用例数据。"""
    excel_path = Path(path)
    if not excel_path.exists():
        raise FileNotFoundError(f"测试数据文件不存在: {excel_path}")

    workbook = load_workbook(excel_path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        header_row = next(rows, None)
        if not header_row:
            return []

        headers = [
            str(header).strip() if header is not None else f"column_{index}"
            for index, header in enumerate(header_row)
        ]
        cases: list[dict[str, Any]] = []
        for row in rows:
            if _is_blank_row(row or ()):
                continue
            cases.append(
                {
                    header: _normalize_value(
                        header,
                        row[index] if index < len(row) else None,
                    )
                    for index, header in enumerate(headers)
                }
            )
        return cases
    finally:
        workbook.close()


def load_cases(path: str | Path) -> list[dict[str, Any]]:
    """按文件后缀读取测试用例，支持 .csv 和 .xlsx。"""
    case_path = Path(path)
    suffix = case_path.suffix.lower()
    if suffix == ".csv":
        return read_csv_cases(case_path)
    if suffix in {".xlsx", ".xlsm"}:
        return read_excel_cases(case_path)
    raise ValueError(
        f"不支持的测试数据文件格式: {suffix}，仅支持 .csv / .xlsx"
    )


def enabled_cases(cases: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """过滤 execute 列标记为禁用的用例；未设置时默认启用。"""
    return [
        case
        for case in cases
        if parse_bool(case.get("execute"), default=True)
    ]


def resolve_cases(
    code_cases: list[dict[str, Any]],
    file_path: str | Path,
    source: str | None = None,
) -> list[dict[str, Any]]:
    """按数据源模式返回用例：code 用代码，file 用文件，auto 优先文件。"""
    mode = (source or os.getenv("TEST_DATA_SOURCE", "auto")).strip().lower()
    if mode not in SUPPORTED_DATA_SOURCES:
        raise ValueError(
            f"不支持的数据源模式: {mode}，可选值: {', '.join(sorted(SUPPORTED_DATA_SOURCES))}"
        )

    if mode == "code":
        return list(code_cases)
    if mode == "file":
        return enabled_cases(load_cases(file_path))
    if Path(file_path).exists():
        return enabled_cases(load_cases(file_path))
    return list(code_cases)
