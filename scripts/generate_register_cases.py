"""生成注册测试用例 CSV 和 xlsx。"""

import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from data.register_data import DEFAULT_PASSWORD, REGISTER_NEGATIVE_CASES


PROJECT_ROOT = Path(__file__).resolve().parent.parent
CSV_PATH = PROJECT_ROOT / "注册测试用例_最终.csv"
XLSX_PATH = PROJECT_ROOT / "注册测试用例_最终.xlsx"

HEADERS = [
    "用例编号",
    "用例名称",
    "模块",
    "测试类型",
    "设计方法",
    "优先级",
    "前置条件",
    "测试数据",
    "测试步骤",
    "预期结果",
    "需求追踪",
    "执行状态",
    "备注",
]

FIELD_LABELS = {
    "email": "邮箱",
    "mobile": "手机号",
    "password": "密码",
    "re_password": "确认密码",
    "first_name": "名字",
    "last_name": "姓氏",
    "agreement": "协议勾选",
}

_DEFAULT_REGISTER_FIELDS = {
    "email": "RANDOM",
    "mobile": "RANDOM",
    "password": DEFAULT_PASSWORD,
    "re_password": DEFAULT_PASSWORD,
    "first_name": "自动化",
    "last_name": "测试",
    "agreement": True,
}


def _field_overrides(case):
    """把外部用例行中偏离默认值的字段提取为可读覆盖项。"""
    overrides = {}
    for key, default in _DEFAULT_REGISTER_FIELDS.items():
        value = case.get(key, default)
        if value != default:
            overrides[key] = value
    return overrides


def _test_data_text(case):
    """把反例 overrides 转为可读测试数据。"""
    overrides = _field_overrides(case)
    if not overrides:
        return "邮箱/手机号随机生成；密码 Test123456；姓名：自动化测试；勾选协议"
    parts = []
    for key, value in overrides.items():
        label = FIELD_LABELS.get(key, key)
        if key == "agreement":
            value = "不勾选"
        elif value == "":
            value = "空"
        parts.append(f"{label}: {value}")
    return "；".join(parts)


def _expected_result(case):
    """根据错误位置生成预期结果。"""
    if case.get("error_field") == "modal":
        return f"停留在注册页并弹出：{case['expected_keyword']}"
    return f"停留在注册页并显示：{case['expected_keyword']}"


def _negative_steps(case):
    """根据覆盖字段生成测试步骤。"""
    field = next(iter(_field_overrides(case)), "数据")
    label = FIELD_LABELS.get(field, field)
    return (
        "1. 打开注册页；"
        f"2. 填写有效基础字段；3. 按测试数据设置{label}；"
        "4. 点击立即登记；5. 校验页面提示"
    )


def build_rows():
    """构建 CSV/xlsx 数据行。"""
    rows = [
        {
            "用例编号": "REG-001",
            "用例名称": "随机邮箱和手机号正常注册",
            "模块": "注册",
            "测试类型": "功能测试",
            "设计方法": "场景法",
            "优先级": "高",
            "前置条件": "已打开注册页，浏览器可用",
            "测试数据": "邮箱/手机号由 Faker 随机生成，密码 Test123456，姓名：自动化测试，勾选协议",
            "测试步骤": (
                "1. 打开注册页；2. 选择中国区号；"
                "3. 填写随机邮箱和手机号；4. 填写密码、确认密码、姓名；"
                "5. 勾选协议；6. 点击立即登记"
            ),
            "预期结果": "注册成功并进入会员中心页面",
            "需求追踪": "注册功能",
            "执行状态": "未执行",
            "备注": "邮箱和手机号每次执行随机生成",
        },
    ]

    for index, case in enumerate(REGISTER_NEGATIVE_CASES, start=2):
        field = next(iter(_field_overrides(case)), "数据")
        rows.append(
            {
                "用例编号": f"REG-{index:03d}",
                "用例名称": case["case"],
                "模块": "注册",
                "测试类型": "功能测试",
                "设计方法": case["design_method"],
                "优先级": "中",
                "前置条件": "已打开注册页，浏览器可用",
                "测试数据": _test_data_text(case),
                "测试步骤": _negative_steps(case),
                "预期结果": _expected_result(case),
                "需求追踪": "注册功能",
                "执行状态": "未执行",
                "备注": f"校验字段：{FIELD_LABELS.get(field, field)}",
            }
        )
    return rows


def write_csv(rows):
    """写入 CSV，使用 utf-8-sig 便于 Excel 打开。"""
    with CSV_PATH.open("w", encoding="utf-8-sig", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=HEADERS)
        writer.writeheader()
        writer.writerows(rows)


def write_xlsx(rows):
    """写入 xlsx，表头黑色文字、不加粗、行高 32。"""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "注册用例"
    worksheet.append(HEADERS)
    for row in rows:
        worksheet.append([row[header] for header in HEADERS])

    for cell in worksheet[1]:
        cell.font = Font(color="000000", bold=False)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    worksheet.row_dimensions[1].height = 32

    for column_index, header in enumerate(HEADERS, start=1):
        max_length = len(header)
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            value = row[column_index - 1]
            max_length = max(max_length, len(str(value or "")))
        worksheet.column_dimensions[get_column_letter(column_index)].width = min(
            max_length + 4, 80
        )

    for row in worksheet.iter_rows(min_row=2):
        max_length = max(len(str(cell.value or "")) for cell in row)
        worksheet.row_dimensions[row[0].row].height = max(32, min(80, max_length * 0.8))
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="center")

    workbook.save(XLSX_PATH)


def main():
    """生成两个最终用例文件。"""
    rows = build_rows()
    write_csv(rows)
    write_xlsx(rows)
    print(f"CSV: {CSV_PATH}")
    print(f"XLSX: {XLSX_PATH}")
    print(f"用例数: {len(rows)}")


if __name__ == "__main__":
    main()
