"""生成核心业务测试用例 CSV 和 xlsx。"""

import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


PROJECT_ROOT = Path(__file__).resolve().parent.parent
CSV_PATH = PROJECT_ROOT / "核心业务测试用例_最终.csv"
XLSX_PATH = PROJECT_ROOT / "核心业务测试用例_最终.xlsx"

HEADERS = [
    "用例编号",
    "用例名称",
    "模块",
    "测试类型",
    "设计方法",
    "优先级",
    "前置条件",
    "测试数据",
    "测试步骤",
    "预期结果",
    "需求追踪",
    "执行状态",
    "备注",
]


def build_rows():
    """构建核心业务用例行。"""
    return [
        {
            "用例编号": "CORE-SEARCH-001",
            "用例名称": "核心关键词搜索商品",
            "模块": "商品搜索",
            "测试类型": "功能测试",
            "设计方法": "场景法",
            "优先级": "P0",
            "前置条件": "已打开公开首页",
            "测试数据": "关键词：laser",
            "测试步骤": "1. 在顶部搜索框输入 laser；2. 按回车；3. 校验搜索结果页",
            "预期结果": "URL 包含 action=search&keyword=laser，页面展示至少一条激光相关商品",
            "需求追踪": "商城商品检索",
            "执行状态": "已自动化",
            "备注": "AI 赋能生成；对应 test_search_returns_laser_products",
        },
        {
            "用例编号": "CORE-STORE-001",
            "用例名称": "商品详情展示价格与库存",
            "模块": "商品详情",
            "测试类型": "功能测试",
            "设计方法": "等价类划分",
            "优先级": "P0",
            "前置条件": "已打开测试商品详情页",
            "测试数据": "商品：Laser hair removal - Brazilian",
            "测试步骤": "1. 打开商品详情页；2. 校验商品标题、价格、库存状态",
            "预期结果": "商品标题正确，单价显示 234，库存状态为 In Stock",
            "需求追踪": "商城商品展示",
            "执行状态": "部分自动化",
            "备注": "AI 赋能生成；价格断言已自动化，库存文案由页面稳定元素承载",
        },
        {
            "用例编号": "CORE-STORE-002",
            "用例名称": "商品加入购物车且小计正确",
            "模块": "购物车",
            "测试类型": "功能测试",
            "设计方法": "场景法",
            "优先级": "P0",
            "前置条件": "已打开测试商品详情页，购物车为空",
            "测试数据": "商品单价：234；数量：1",
            "测试步骤": "1. 点击 Add to cart；2. 等待购物车商品出现；3. 校验小计",
            "预期结果": "购物车包含 Laser hair removal - Brazilian，小计显示 $ 234",
            "需求追踪": "商城加购",
            "执行状态": "已自动化",
            "备注": "AI 赋能生成；对应 test_product_detail_add_to_cart",
        },
        {
            "用例编号": "CORE-STORE-003",
            "用例名称": "会员商品未登录结算跳转登录",
            "模块": "结算",
            "测试类型": "功能测试",
            "设计方法": "状态迁移",
            "优先级": "P0",
            "前置条件": "未登录会员账号，购物车已加入会员专属商品",
            "测试数据": "无",
            "测试步骤": "1. 点击 Proceed Checkout；2. 等待跳转；3. 校验登录页",
            "预期结果": "跳转到 /account/login，提示会员商品需先登录或注册",
            "需求追踪": "商城会员结算",
            "执行状态": "已自动化",
            "备注": "AI 赋能生成；对应 test_member_checkout_requires_login",
        },
        {
            "用例编号": "CORE-BOOK-001",
            "用例名称": "未选择服务不可进入下一步",
            "模块": "预约",
            "测试类型": "功能测试",
            "设计方法": "判定表",
            "优先级": "P0",
            "前置条件": "已打开免费咨询预约页",
            "测试数据": "服务选择：空",
            "测试步骤": "1. 不选择服务；2. 点击 Next；3. 校验当前步骤",
            "预期结果": "停留在服务选择步骤，页面不出现 Proceed to Confirm Booking",
            "需求追踪": "预约服务选择",
            "执行状态": "已自动化",
            "备注": "AI 赋能生成；对应 test_booking_next_without_service_keeps_service_step",
        },
        {
            "用例编号": "CORE-BOOK-002",
            "用例名称": "选择服务后进入日期时间步骤",
            "模块": "预约",
            "测试类型": "功能测试",
            "设计方法": "场景法",
            "优先级": "P0",
            "前置条件": "已打开免费咨询预约页",
            "测试数据": "选择第一个服务",
            "测试步骤": "1. 选择第一个服务；2. 点击 Next；3. 校验日期时间步骤",
            "预期结果": "页面显示 Day and time 和 Proceed to Confirm Booking",
            "需求追踪": "预约流程",
            "执行状态": "已自动化",
            "备注": "AI 赋能生成；对应 test_select_service_opens_day_time_step",
        },
        {
            "用例编号": "CORE-STORE-004",
            "用例名称": "修改商品数量后小计正确",
            "模块": "购物车",
            "测试类型": "功能测试",
            "设计方法": "边界值分析",
            "优先级": "P1",
            "前置条件": "已打开测试商品详情页，购物车为空",
            "测试数据": "数量：2；单价：234",
            "测试步骤": "1. 将数量改为 2；2. 点击 Add to cart；3. 校验小计",
            "预期结果": "购物车小计显示 $ 468",
            "需求追踪": "商城加购",
            "执行状态": "待自动化",
            "备注": "AI 赋能生成；数量输入框存在 data-min=1，可作为下一步补充",
        },
        {
            "用例编号": "CORE-BOOK-003",
            "用例名称": "未选择日期时间不可确认预约",
            "模块": "预约",
            "测试类型": "功能测试",
            "设计方法": "判定表",
            "优先级": "P1",
            "前置条件": "已选择服务并进入日期时间步骤",
            "测试数据": "日期时间：空",
            "测试步骤": "1. 进入日期时间步骤；2. 不选择日期时间；3. 点击 Proceed to Confirm Booking；4. 校验提示",
            "预期结果": "页面阻止确认预约并显示必选提示",
            "需求追踪": "预约流程",
            "执行状态": "待自动化",
            "备注": "AI 赋能生成；已自动化到日期时间步骤，真实提交由 booking_submit 标记控制",
        },
        {
            "用例编号": "CORE-BOOK-003",
            "用例名称": "选择日期时间进入预约确认",
            "模块": "预约",
            "测试类型": "功能测试",
            "设计方法": "场景法",
            "优先级": "P0",
            "前置条件": "已登录会员账号并选择服务",
            "测试数据": "日期：2026-08-18；时间：09:30",
            "测试步骤": "1. 选择服务；2. 进入日期时间步骤；3. 选择日期和时间；4. 点击前往确认",
            "预期结果": "结算页展示所选服务、预约时间，并进入提交预约步骤",
            "需求追踪": "预约确认",
            "执行状态": "已自动化",
            "备注": "AI 赋能生成；对应 test_booking_confirmation_page_shows_service_and_time",
        },
        {
            "用例编号": "CORE-BOOK-004",
            "用例名称": "继续添加服务",
            "模块": "预约",
            "测试类型": "功能测试",
            "设计方法": "场景法",
            "优先级": "P1",
            "前置条件": "已登录会员账号，已选择服务并进入日期时间步骤",
            "测试数据": "第一个服务：Aesthetic consultation",
            "测试步骤": "1. 选择服务并进入日期时间；2. 选择日期时间；3. 点击继续添加服务；4. 回到服务选择并选择第二个服务",
            "预期结果": "页面回到服务选择步骤，可继续选择第二个服务，日期时间选择不被丢失",
            "需求追踪": "预约多服务",
            "执行状态": "部分自动化",
            "备注": "AI 赋能生成；对应 test_add_another_service_returns_to_service_step",
        },
        {
            "用例编号": "CORE-BOOK-005",
            "用例名称": "重复提交同一预约时段",
            "模块": "预约",
            "测试类型": "功能测试",
            "设计方法": "错误猜测",
            "优先级": "P1",
            "前置条件": "BOOKING_SUBMIT_ALLOWED=true，测试环境允许真实预约",
            "测试数据": "同一服务、同一日期时间",
            "测试步骤": "1. 提交第一次预约；2. 再次选择相同服务和时间提交；3. 校验系统拦截或提示重复",
            "预期结果": "系统阻止重复预约，或明确提示该时段已预约",
            "需求追踪": "预约重复提交",
            "执行状态": "条件执行",
            "备注": "AI 赋能生成；真实提交需显式开启 BOOKING_SUBMIT_ALLOWED",
        },
        {
            "用例编号": "CORE-BOOK-006",
            "用例名称": "结算页 x 取消预约服务",
            "模块": "预约",
            "测试类型": "功能测试",
            "设计方法": "状态迁移",
            "优先级": "P1",
            "前置条件": "已登录并进入预约确认/结算页",
            "测试数据": "无",
            "测试步骤": "1. 进入预约确认/结算页；2. 点击服务行 x；3. 校验服务被移除",
            "预期结果": "该预约服务从结算页移除，URL 出现 removeIds",
            "需求追踪": "预约取消",
            "执行状态": "已自动化",
            "备注": "AI 赋能生成；对应 test_booking_cancel_service_from_checkout_x",
        },
        {
            "用例编号": "CORE-PAY-001",
            "用例名称": "购物车金额与应付金额一致",
            "模块": "支付",
            "测试类型": "功能测试",
            "设计方法": "等价类划分",
            "优先级": "P0",
            "前置条件": "已登录会员账号",
            "测试数据": "商品：Laser hair removal - Brazilian；金额：234",
            "测试步骤": "1. 商品加入购物车；2. 进入结算；3. 校验应付金额",
            "预期结果": "结算页应付金额显示 $ 234，与购物车小计一致",
            "需求追踪": "金额计算",
            "执行状态": "已自动化",
            "备注": "AI 赋能生成；对应 test_checkout_amount_calculation",
        },
        {
            "用例编号": "CORE-PAY-002",
            "用例名称": "当前 Web 仅 ATM 支付",
            "模块": "支付",
            "测试类型": "功能测试",
            "设计方法": "等价类划分",
            "优先级": "P0",
            "前置条件": "已登录会员账号并进入结算页",
            "测试数据": "支付方式候选",
            "测试步骤": "1. 进入结算页；2. 列出支付方式；3. 校验 ATM 为唯一选项",
            "预期结果": "支付方式仅 ATM，符合到店支付配置",
            "需求追踪": "支付方式",
            "执行状态": "已自动化",
            "备注": "AI 赋能生成；对应 test_checkout_payment_method_is_atm_only",
        },
        {
            "用例编号": "CORE-PAY-003",
            "用例名称": "支付成功",
            "模块": "支付",
            "测试类型": "功能测试",
            "设计方法": "场景法",
            "优先级": "P1",
            "前置条件": "已接入 CMS 支付沙箱或模拟网关",
            "测试数据": "使用测试支付凭据",
            "测试步骤": "1. 提交订单；2. 完成 ATM/模拟支付；3. 校验订单状态为已支付",
            "预期结果": "支付成功，订单状态更新为已支付",
            "需求追踪": "支付成功",
            "执行状态": "环境受限",
            "备注": "AI 赋能生成；当前站点无 CMS，无法端到端执行",
        },
        {
            "用例编号": "CORE-PAY-004",
            "用例名称": "支付失败",
            "模块": "支付",
            "测试类型": "功能测试",
            "设计方法": "错误猜测",
            "优先级": "P1",
            "前置条件": "已接入 CMS 支付沙箱或模拟网关",
            "测试数据": "错误支付凭据/余额不足",
            "测试步骤": "1. 提交订单；2. 模拟支付失败；3. 校验订单保持待付款",
            "预期结果": "支付失败，订单不更新为已支付，页面提示失败",
            "需求追踪": "支付失败",
            "执行状态": "环境受限",
            "备注": "AI 赋能生成；当前站点无 CMS，无法端到端执行",
        },
    ]


def write_csv(rows):
    """写入 CSV，使用 utf-8-sig 便于 Excel 打开。"""
    with CSV_PATH.open("w", encoding="utf-8-sig", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=HEADERS)
        writer.writeheader()
        writer.writerows(rows)


def write_xlsx(rows):
    """写入 xlsx，表头黑色文字、不加粗、行高 32。"""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "核心业务用例"
    worksheet.append(HEADERS)
    for row in rows:
        worksheet.append([row[header] for header in HEADERS])

    for cell in worksheet[1]:
        cell.font = Font(color="000000", bold=False)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    worksheet.row_dimensions[1].height = 32

    for column_index, header in enumerate(HEADERS, start=1):
        max_length = len(header)
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            value = row[column_index - 1]
            max_length = max(max_length, len(str(value or "")))
        worksheet.column_dimensions[get_column_letter(column_index)].width = min(
            max_length + 4, 80
        )

    for row in worksheet.iter_rows(min_row=2):
        max_length = max(len(str(cell.value or "")) for cell in row)
        worksheet.row_dimensions[row[0].row].height = max(32, min(80, max_length * 0.8))
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="center")

    workbook.save(XLSX_PATH)


def main():
    """生成核心业务最终用例文件。"""
    rows = build_rows()
    write_csv(rows)
    write_xlsx(rows)
    print(f"CSV: {CSV_PATH}")
    print(f"XLSX: {XLSX_PATH}")
    print(f"用例数: {len(rows)}")


if __name__ == "__main__":
    main()
