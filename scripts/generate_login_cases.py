"""生成登录测试用例最终 CSV 和 xlsx。"""

import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from config import CASE_FILES_DIR
from utils.data_reader import load_cases


PROJECT_ROOT = Path(__file__).resolve().parent.parent
CSV_PATH = PROJECT_ROOT / "登录测试用例_最终.csv"
XLSX_PATH = PROJECT_ROOT / "登录测试用例_最终.xlsx"
LOGIN_CASES_FILE = CASE_FILES_DIR / "login_cases.csv"

HEADERS = [
    "用例编号",
    "用例名称",
    "模块",
    "测试类型",
    "设计方法",
    "优先级",
    "前置条件",
    "测试数据",
    "测试步骤",
    "预期结果",
    "需求追踪",
    "执行状态",
    "备注",
]


def build_rows():
    """把数据驱动登录用例转换为最终交付格式。"""
    cases = load_cases(LOGIN_CASES_FILE)
    rows = []
    design_methods = [
        "状态迁移",
        "错误猜测",
        "场景法",
        "边界值分析",
        "边界值分析",
        "错误猜测",
        "判定表",
        "边界值分析",
        "错误猜测",
        "等价类划分",
    ]
    for index, case in enumerate(cases, start=1):
        success = case.get("expected_result") == "success"
        rows.append(
            {
                "用例编号": f"LOGIN-{index:03d}",
                "用例名称": case.get("case", ""),
                "模块": "登录",
                "测试类型": "功能测试",
                "设计方法": design_methods[min(index - 1, len(design_methods) - 1)],
                "优先级": "高" if success else "中",
                "前置条件": "已打开登录页，浏览器可用",
                "测试数据": f"账号：{case.get('username', '')}；密码：{case.get('password', '空')}",
                "测试步骤": (
                    "1. 打开登录页；2. 输入账号；3. 输入密码；"
                    "4. 点击 Sign-in；5. 校验登录结果"
                ),
                "预期结果": (
                    "登录成功并离开登录页"
                    if success
                    else "停留在登录页并显示错误提示"
                ),
                "需求追踪": "登录功能",
                "执行状态": "已自动化" if case.get("execute", True) else "待执行",
                "备注": "AI 赋能生成；数据驱动源为 data/cases/login_cases.csv",
            }
        )
    return rows


def write_csv(rows):
    """写入 CSV，使用 utf-8-sig 便于 Excel 打开。"""
    with CSV_PATH.open("w", encoding="utf-8-sig", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=HEADERS)
        writer.writeheader()
        writer.writerows(rows)


def write_xlsx(rows):
    """写入 xlsx，表头黑色文字、不加粗、行高 32。"""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "登录用例"
    worksheet.append(HEADERS)
    for row in rows:
        worksheet.append([row[header] for header in HEADERS])

    for cell in worksheet[1]:
        cell.font = Font(color="000000", bold=False)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    worksheet.row_dimensions[1].height = 32

    for column_index, header in enumerate(HEADERS, start=1):
        max_length = len(header)
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            value = row[column_index - 1]
            max_length = max(max_length, len(str(value or "")))
        worksheet.column_dimensions[get_column_letter(column_index)].width = min(
            max_length + 4, 80
        )

    for row in worksheet.iter_rows(min_row=2):
        max_length = max(len(str(cell.value or "")) for cell in row)
        worksheet.row_dimensions[row[0].row].height = max(32, min(80, max_length * 0.8))
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="center")

    workbook.save(XLSX_PATH)


def main():
    """生成最终登录用例文件。"""
    rows = build_rows()
    write_csv(rows)
    write_xlsx(rows)
    print(f"CSV: {CSV_PATH}")
    print(f"XLSX: {XLSX_PATH}")
    print(f"用例数: {len(rows)}")


if __name__ == "__main__":
    main()
